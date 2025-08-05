from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XImage
from PIL import Image as PImage
from datetime import datetime
from pathlib import Path
from io import BytesIO
import numpy as np
import logging

logger = logging.getLogger(__name__)

in2dpi = lambda inch, dpi : (int(inch * dpi), int(inch * dpi))

def plt2npa(fig):
    buffer = BytesIO()
    fig.savefig(buffer, format='png', bbox_inches='tight', pad_inches=0)
    buffer.seek(0)
    image = PImage.open(buffer)
    return np.array(image)

def npa2pyxl(img_array, size):
    # Return a scaled Openpyxl Image from a numpy array
    pil_image = PImage.fromarray(img_array) # Convert numpy array to Pillow image
    pil_image.thumbnail(size) # Resize image while maintaining aspect ratio
    buffer = BytesIO() # Create in-memory buffer for saving image
    pil_image.save(buffer, format='PNG') # Save Pillow image in buffer
    buffer.seek(0) # Reset buffer to beginning
    pyxl_image = XImage(buffer) # Open buffer as Openpyxl image
    return pyxl_image

def create_unique_sheet(workbook, base_title):
    sheet_title = base_title
    counter = 1

    while sheet_title in workbook.sheetnames:
        sheet_title = f"{base_title} ({counter})"
        counter += 1
    
    new_sheet = workbook.create_sheet(f"{sheet_title}")
    return new_sheet

def add_row(ws, start_row, start_col, data, num_format, font_style=None, border_style=None, align_style=None ):
    for index, value in enumerate(data):
        cell = ws.cell(row=start_row, column=start_col+index, value=value)
        cell.font = font_style
        cell.border = border_style
        cell.alignment = align_style
        cell.number_format = num_format

def OutputImageData(data_dir, mean_rgb, mean_hsv, img_conf, img_hist, img_blob):
    # Create a path for the Excel file based on the current date
    path = Path(f"{data_dir}/{datetime.now().strftime("%Y%m%d")}_ImageData.xlsx")
    curtime = datetime.now().strftime("%I|%M%p")

    # Open or create the Excel file
    if not path.is_file():
        wb = Workbook()
        ws = wb.active
        ws.title = f"{curtime}"
    else:
        wb = load_workbook(filename=path)
        ws = create_unique_sheet(wb, f"{curtime}")

    # Write data to the Excel file
    border_thin = Side(style="thin")
    border_thin_outline = Border(top=border_thin, bottom=border_thin, right=border_thin, left=border_thin)
    align_centered = Alignment(horizontal="center", vertical="center")

    ws["B2"] = "Average HSV"
    add_row(ws, 3, 2, ("H", "S", "V", "S%", "V%"), "", Font(bold=True), border_thin_outline, align_centered)
    add_row(ws, 4, 2, mean_hsv, "0.0", Font(bold=False), border_thin_outline, align_centered)
    add_row(ws, 4, 5, ('=C4/255', '=D4/255'), "0.0%", Font(bold=False), border_thin_outline, align_centered)

    ws["H2"] = "Average RGB"
    add_row(ws, 3, 8, ("R", "G", "B"), "", Font(bold=True), border_thin_outline, align_centered)
    add_row(ws, 4, 8, mean_rgb, "0", Font(bold=False), border_thin_outline, align_centered)

    ws.add_image(npa2pyxl(img_conf, in2dpi(5, 96)), "N2") # Add confirmation image to worksheet
    ws.add_image(npa2pyxl(img_hist, in2dpi(7.5, 96)), "B6") # Add histogram image to worksheet
    ws.add_image(npa2pyxl(img_blob, in2dpi(0.75, 96)), "L2") # Add color blob image to worksheet

    # Save the workbook
    try:
        wb.save(path)
    except PermissionError as e:
        logger.error(f"Permission denied: {e}. Please close the file if it's open in another application.")
    else:
        logger.info(f"Data successfully written to {path}.")
    finally:
        wb.close()
        logger.info("Workbook closed.")

def OutputAudioData(data_dir, img_waveform, img_spectrum, img_fft, img_env, mag, freq, peak_amp, rms_amp):
    # Create a path for the Excel file based on the current date
    path = Path(f"{data_dir}/{datetime.now().strftime("%Y%m%d")}_AudioData.xlsx")
    curtime = datetime.now().strftime("%I|%M%p")

    # Open or create the Excel file
    if not path.is_file():
        wb = Workbook()
        ws = wb.active
        ws.title = f"{curtime}"
    else:
        wb = load_workbook(filename=path)
        ws = create_unique_sheet(wb, f"{curtime}")

    # Write data to the Excel file
    ws.add_image(npa2pyxl(img_waveform, in2dpi(8, 96)), "B2") # Add Waveform image to worksheet
    ws.add_image(npa2pyxl(img_spectrum, in2dpi(8, 96)), "B16") # Add Spectrogram image to worksheet
    ws.add_image(npa2pyxl(img_fft, in2dpi(8, 96)), "O2") # Add FFT image to worksheet
    ws.add_image(npa2pyxl(img_env, in2dpi(8, 96)), "O18") # Add Envelope image to worksheet

    cell = ws.cell(15, 16, "Max magnitude:")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(15, 17, mag)
    cell.number_format = "0.0"
    cell = ws.cell(16, 16, "Frequency (Hz):")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(16, 17, freq)
    cell.number_format = "0.0"

    cell = ws.cell(31, 16, "Peak Amplitude (db):")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(31, 17, peak_amp)
    cell.number_format = "0.0"
    cell = ws.cell(32, 16, "RMS Amplitude (db):")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell = ws.cell(32, 17, rms_amp)
    cell.number_format = "0.0"


    # Save the workbook
    try:
        wb.save(path)
    except PermissionError as e:
        logger.error(f"Permission denied: {e}. Please close the file if it's open in another application.")
    else:
        logger.info(f"Data successfully written to {path}.")
    finally:
        wb.close()
        logger.info("Workbook closed.")
